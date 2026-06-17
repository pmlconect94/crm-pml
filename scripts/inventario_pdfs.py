"""
Genera un Excel inventario de los contratos Blufin para saber qué PDFs juntar
(contrato + factura). Lee de Supabase vía REST (anon key + dev_open) y extrae el
número de factura (C####) y el cliente de las observaciones.

Salida: uploads/inventario-blufin-pdfs.xlsx

Uso: python scripts/inventario_pdfs.py
"""
import json
import re
import urllib.request
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

FACTURA_RE = re.compile(r"(C\d{3,})")
CLIENTE_RE = re.compile(r"Cliente\s+([^·]+)")


def load_env(path=".env.local"):
    env = {}
    for line in Path(path).read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if "=" in line and not line.startswith("#"):
            k, v = line.split("=", 1)
            env[k.strip()] = v.strip()
    return env


def main():
    env = load_env()
    url = env["VITE_SUPABASE_URL"].rstrip("/")
    key = env["VITE_SUPABASE_ANON_KEY"]

    cols = "folio,fecha,observaciones,total_usd,total_kg,contrato_pdf_path,factura_pdf_path"
    req = urllib.request.Request(
        f"{url}/rest/v1/blufin_contratos?select={cols}&order=folio.asc",
        headers={"apikey": key, "Authorization": f"Bearer {key}", "Accept-Profile": "crm"},
    )
    rows = json.loads(urllib.request.urlopen(req).read())

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario PDFs"
    headers = ["Folio", "Fecha", "# Factura", "Cliente", "Total USD", "Total Kg",
               "Tiene PDF contrato", "Tiene PDF factura"]
    ws.append(headers)
    bold = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="0A2540")
    for c in ws[1]:
        c.font = bold
        c.fill = fill

    sin_factura = 0
    for r in rows:
        obs = r.get("observaciones") or ""
        mf = FACTURA_RE.search(obs)
        mc = CLIENTE_RE.search(obs)
        factura = mf.group(1) if mf else ""
        if not factura:
            sin_factura += 1
        ws.append([
            r.get("folio"),
            r.get("fecha"),
            factura,
            (mc.group(1).strip() if mc else ""),
            float(r["total_usd"]) if r.get("total_usd") else None,
            float(r["total_kg"]) if r.get("total_kg") else None,
            "sí" if r.get("contrato_pdf_path") else "—",
            "sí" if r.get("factura_pdf_path") else "—",
        ])

    widths = [16, 12, 12, 22, 14, 14, 18, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    out = Path("uploads/inventario-blufin-pdfs.xlsx")
    out.parent.mkdir(exist_ok=True)
    wb.save(out)
    print(json.dumps({
        "archivo": str(out),
        "contratos": len(rows),
        "con_factura": len(rows) - sin_factura,
        "sin_factura": sin_factura,
    }, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
